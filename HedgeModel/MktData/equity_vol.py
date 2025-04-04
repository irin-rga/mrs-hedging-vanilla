from datetime import datetime, date
from openpyxl import load_workbook
from pathlib import Path
import pandas as pd
import typing  # noqa: N999
import math

from crossasset.helpers.matrix import Matrix as MatrixHelper
from crossasset.helpers.data import Data as DataHelper
from crossasset.market_data.dividend import Dividend
from crossasset.market_data.eq_vol import EqVol
from crossasset.helpers.matrix import Matrix
from crossasset.market_data.yld import Yld

from nxpy import pro


class EquityVol:

    """USD Markit Equity ImpliedVol Elements"""

    __currency = "USD"    
    # __vol_path = '\\\\rgare.net\\stlfinmkts\\MarketRiskMgmt\\Market Hedge VA\\Markit Vol Surface'
    __vol_path = r'\\rgare.net\stlfinmkts\MarketRiskMgmt\Market Hedge VA\Markit Vol Surface'
    __vol_file_root_name = 'Markit_Vol_Surface_'
    __vol_file_ext = '.xlsx'

    __vol_tab_lu = {
        'MXEA Index': 'EAFE',
        'SPX Index' : 'S&P500',
        'NDX Index' : 'NASDAQ',
        'SPMARC5P Index' : 'S&P_MARC_5%',
        'RTY Index' : 'RUSSELL2000'
    }

    __sheet_rng_lu = {
        'All_Dts' : 'A2:A14',
        'Expiry_Dts' : 'A3:A14',
        'Strikes' : 'B2:N2',
        'Full_IV_Matrix' : 'A2:N14',
        'IV_Data' : 'B3:N14',
        'YC_DFs' : 'S3:S14',
        'Dvd_Rts' : 'R3:R14'
    }
    # __vol_path = Path(Path(__file__).parents[2], "Market_Data", "EQ_Market_Data", "SP500_MoneynessVol.csv")

    def __init__(
        self,
        app: pro.Application,
        market_reference_date: typing.Union[str, datetime.date],  # noqa: FA100
        index_name: str,
        spot_px: float,
        # yield_curve: str,
        container: typing.Optional[str] = None,  # noqa: FA100
    ):
        """set market reference date, yield curve and container, then dividend yield curve, vol surface"""
        # set market reference date and container
        if isinstance(market_reference_date, date):
            self.now_date = market_reference_date
        else:
            self.now_date = datetime.strptime(market_reference_date, '%Y%m%d').date()


        self.app = app
        self.yc_data = None
        self.bbg_idx = index_name
        self.index_name = index_name.replace(' Index', '')        
        self.container = container
        
        self.id_prefix = self.index_name + '_' + self.now_date.strftime('%Y%m%d')


        # for key, val in self.__sheet_rng_lu.items():
        #     print(key, val)

        # Read raw data pieces from excel using pyxl
        self.excel_data = self.read_all_excel_data()

        # Read raw data pieces from excel using pandas
        # self.excel_data = self.read_all_vol_data_from_excel()
        
        # build MDEs
        self.spot_price = spot_px
        self.yield_curve = self.build_yield_curve()
        self.div_curve = self.build_dividend_curve()
        self.vol_surf = self.build_vol_surface()

    @property
    def vol_file(self):
        return Path(self.__vol_path, self.__vol_file_root_name + self.now_date.strftime('%Y%m%d') + self.__vol_file_ext)
    
    @property
    def sheet_name(self):
        return self.__vol_tab_lu[self.bbg_idx]
    
    def sheet_rng(self, data_type: str):
        return self.__sheet_rng_lu[data_type]
    
    def get_excel_data_for_datatype(self, data_type):
        """Reads data from a specified Excel sheet and range."""
        range_str = self.sheet_rng(data_type)
        df = pd.read_excel(self.vol_file, sheet_name=self.sheet_name, usecols=range_str)
        return df

    def read_all_vol_data_from_excel(self):        
        """Reads data from all excel ranges into memory"""
        
        excel_data = {}

        for data_type in self.__sheet_rng_lu.keys():
            excel_data[data_type] = self.get_excel_data_for_datatype(data_type)
        
        return excel_data

    def read_all_excel_data(self):

        # print(f'Reading Excel Data for {self.index_name} on {self.now_date}')

        excel_data = {}

        wb = load_workbook(filename=self.vol_file, read_only=True)
        ws = wb[self.sheet_name]

        def read_1d_rng(rng: str):
            # Extract data into 1d
            data = []
            for row in ws[rng]:    
                for cell in row:
                    data.append(cell.value)
            return data
        
        def read_2d_rng(rng: str):
            data = []
            for row in ws[rng]:
                row_data = []
                for cell in row:
                    row_data.append(cell.value)
                data.append(row_data)
            return data

        for key, rng in self.__sheet_rng_lu.items():
            
            read_xl_rng_fn = read_2d_rng if key in ['Full_IV_Matrix', 'IV_Data'] else read_1d_rng
            excel_data[key] = read_xl_rng_fn(rng)

        # Dates don't need conversion to datetime, as they already appear to be in datetime format
        # for key, values in excel_data.items():
        #     if key in ['All_Dts', 'Expiry_Dts']:
        #         excel_data[key] = [pd.to_datetime(row, format="%m/%d/%Y") for row in values]

        return excel_data

    def build_yield_curve(self):        
        """build yield curve from Markit Discount Factors"""

        # print(f'Building Yield Curve for {self.index_name} on {self.now_date}')

        disc_fctrs_dates = self.excel_data['All_Dts']
        disc_fctrs = self.excel_data['YC_DFs']
        disc_fctrs.insert(0, 1.0000)
        
        yc_data = DataHelper.from_dict({"DATE": disc_fctrs_dates, "DISCOUNTFACTOR": disc_fctrs}, app=self.app, custom_id=self.id_prefix+'_YCData')
        yc = Yld.discount_factors(app=self.app, interp="CUBIC SPLINE", currency="USD", data=yc_data, basis="Act/365.25", container=self.container, custom_id=self.id_prefix+'_YC')

        self.yc_data = yc_data

        return yc


    def build_dividend_curve(self):
        """build equity dividend curve from forward dividend yields rate NOTE:  Markit Dividend Yields are Spot!"""

        # print(f'Building Dividend Curve for {self.index_name} on {self.now_date}')

        # get the dividend data
        # all_dates = self.excel_data['All_Dts']
        div_dates = self.excel_data["Expiry_Dts"]
        div_spot_rts = self.excel_data['Dvd_Rts']

        # Start calculations for converting spot to forward
        start_dt = datetime.combine(self.now_date, datetime.min.time())        
        exp_times = [(dt-start_dt).days/365.25 for dt in div_dates]
        cont_spot_rts = [math.log((1+div_spot_rts[t])**exp_times[t])/exp_times[t] for t in range(len(exp_times))]
        dvd_accum_fctrs = [math.exp(cont_spot_rts[t]*exp_times[t]) for t in range(len(exp_times))]
        dvd_accum_fctrs.insert(0, 1.0)
        chg_in_t = [exp_times[t] if t == 0 else exp_times[t]-exp_times[t-1] for t in range(len(exp_times))]
        div_fwd_rts = [math.log(dvd_accum_fctrs[t+1]/dvd_accum_fctrs[t])/chg_in_t[t] for t in range(len(exp_times))]
        
        div_data = DataHelper.from_dict({"DATE": div_dates, "DIVIDENDYIELD": div_fwd_rts}, app=self.app, custom_id=self.id_prefix+'_DivData')

        cust_id = self.id_prefix + '_DIV'
        
        # build div curve
        div_curve = Dividend.dividend_yields(app=self.app, nowdate=self.now_date, currency="USD", discrete=False, rate_or_dividend=div_data, basis="ACT/365.25", container=self.container, custom_id=cust_id)
        return div_curve

        
        # curve = Dividend.single_continuous_dividend(
        #     nowdate=self.now_date,
        #     currency=self.__currency,
        #     rate_or_dividend=0.0226911211045757,
        #     basis="ACT/365",
        #     asset_name=self.index_name,
        #     container=self._container,
        # )
        # return curve

    def build_vol_surface(self):
        """build equity moneyness volatility surface by reading in a two-dimensional vol matrix"""

        # print(f'Building Vol Surface for {self.index_name} on {self.now_date}')
        
        vol_matrix = Matrix.from_list_of_lists(
            app=self.app,
            matrix_row_headers=self.excel_data["Expiry_Dts"],
            matrix_col_headers=self.excel_data["Strikes"],
            values=self.excel_data["IV_Data"],
            custom_id=self.id_prefix + "_IV_Matrix"
        )

        equity_vol = EqVol.eq_volatility(
            app=self.app,
            nowdate=self.now_date,
            currency=self.__currency,
            volatility_basis="ACT/365.25",
            data=vol_matrix,
            scaling_factor=1.0,
            strike_definition="Relative",
            strike_extrapolation="FLAT",
            time_interpolation_method="Linear",
            time_interpolation_variable="Vol Square T",
            fix_calendar="NONE",
            fix_convention="NONE",
            notice_period="0BD",
            discrete_dividend_convention="STRIKE",
            dividend_curve=self.div_curve,            
            domestic_yield_curve=self.yield_curve,
            spot_price=self.spot_price,
            atm_convention="CALL50",
            strike_interpolation="LINEAR STRIKE",
            container=self.container,
            custom_id=self.id_prefix + "_IVSurf"
        )

        return equity_vol
        

        # Read data
        # vol_df = pd.read_csv(self.__vol_path)
        # vol_df[vol_df.columns[0]] = pd.to_datetime(vol_df[vol_df.columns[0]], format="%m/%d/%Y")
        # vol_df.set_index(vol_df.columns[0], inplace=True)  # noqa: PD002
        # Build vol matrix
        # noinspection PyTypeChecker
        # vol_matrix = Matrix.from_list_of_lists(
        #     matrix_row_headers=vol_df.index.tolist(),
        #     matrix_col_headers=vol_df.columns.tolist(),
        #     values=vol_df.values.tolist(),  # noqa: PD011
        # )
        # build volatility surface by moneyness
        # vol_by_moneyness = EqVol.eq_volatility(
        #     nowdate=self.now_date,
        #     currency=self.__currency,
        #     volatility_basis="ACT/365",
        #     data=vol_matrix,
        #     time_interpolation_method="Linear Spline",
        #     time_extrapolation_method="Flat",
        #     strike_definition="Relative",
        #     fix_calendar="NewYork",
        #     fix_convention="F",
        #     notice_period="2BD",
        #     dividend_curve=self.div_curve,
        #     discrete_dividend_convention="Mixed",
        #     domestic_yield_curve=self._yield_curve,
        #     spot_price=self.spot_price,
        #     container=self._container,
        # )
        # return vol_by_moneyness
